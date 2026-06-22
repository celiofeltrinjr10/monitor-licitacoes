"""
Monitor de Licitações de Obras — PNCP
Interface web Streamlit para o pncp_scraper.py
"""

import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pncp_scraper import (
    consultar_pncp,
    contem_palavra_chave,
    normalizar_linha,
    sanitizar,
    MODALIDADES,
    PALAVRAS_CHAVE_OBRAS,
    CABECALHO,
)

# ── Configuração da página ────────────────────────────────────────────────────
st.set_page_config(
    page_title="Monitor PNCP — Obras",
    page_icon="🏗️",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    [data-testid="stSidebar"] { background-color: #f0f4f8; }
    .block-container { padding-top: 1.5rem; }
    h1 { color: #1F4E78; }
</style>
""", unsafe_allow_html=True)

# ── Cabeçalho ─────────────────────────────────────────────────────────────────
st.title("🏗️ Monitor de Licitações de Obras — PNCP")
st.caption("Infraestrutura · Mobilidade Urbana · Rodovias · Construção Pesada")
st.divider()

# ── Barra lateral — filtros ───────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Parâmetros")

    dias = st.slider(
        "Período de busca (dias atrás)",
        min_value=1, max_value=30, value=3,
        help="Quantos dias retroativos consultar no PNCP"
    )
    if dias > 7:
        st.warning("⚠️ Períodos acima de 7 dias podem causar timeout na nuvem. Prefira até 5 dias.")

    valor_min = st.number_input(
        "Valor mínimo estimado (R$)",
        min_value=0,
        max_value=5_000_000_000,
        value=100_000_000,
        step=50_000_000,
        format="%d",
    )
    st.caption("Contratos com valor R$0 (ex: RDCI / orçamento sigiloso) são sempre incluídos.")

    st.divider()
    st.subheader("Modalidades")
    mods_sel = {}
    for cod, nome in MODALIDADES.items():
        mods_sel[cod] = st.checkbox(nome, value=True, key=f"mod_{cod}")

    st.divider()
    st.subheader("Estados (UF)")
    TODAS_UFS = [
        "AC","AL","AP","AM","BA","CE","DF","ES","GO","MA","MT","MS",
        "MG","PA","PB","PR","PE","PI","RJ","RN","RS","RO","RR","SC","SP","SE","TO"
    ]
    uf_filtro_sidebar = st.multiselect(
        "Filtrar por estado",
        options=TODAS_UFS,
        placeholder="Todos os estados",
        label_visibility="collapsed",
        help="Filtra os resultados após a busca",
    )

    st.divider()
    buscar_btn = st.button(
        "🔍  Buscar Licitações",
        type="primary",
        use_container_width=True,
    )

    with st.expander("📋 Palavras-chave ativas", expanded=False):
        # Inicializa a lista editável no session_state (cópia da lista padrão)
        if "palavras_ativas" not in st.session_state:
            st.session_state.palavras_ativas = list(PALAVRAS_CHAVE_OBRAS)

        # Adicionar nova palavra
        col_input, col_add = st.columns([3, 1])
        with col_input:
            nova = st.text_input("Nova palavra-chave", key="nova_palavra",
                                 label_visibility="collapsed",
                                 placeholder="Ex: obra de drenagem")
        with col_add:
            if st.button("＋ Adicionar", use_container_width=True):
                nova_clean = nova.strip().lower()
                if nova_clean and nova_clean not in st.session_state.palavras_ativas:
                    st.session_state.palavras_ativas.append(nova_clean)
                    st.rerun()

        st.caption(f"Total: {len(st.session_state.palavras_ativas)} palavras-chave")
        st.divider()

        # Lista atual com botão de remover
        for i, p in enumerate(st.session_state.palavras_ativas):
            col_p, col_rem = st.columns([4, 1])
            with col_p:
                st.caption(f"• {p}")
            with col_rem:
                if st.button("✕", key=f"rem_{i}", help=f"Remover '{p}'"):
                    st.session_state.palavras_ativas.pop(i)
                    st.rerun()

        st.divider()
        if st.button("↺ Restaurar padrão", use_container_width=True):
            st.session_state.palavras_ativas = list(PALAVRAS_CHAVE_OBRAS)
            st.rerun()


# ── Estado da sessão ──────────────────────────────────────────────────────────
if "resultados" not in st.session_state:
    st.session_state.resultados = []
if "ultima_busca" not in st.session_state:
    st.session_state.ultima_busca = None
if "log_busca" not in st.session_state:
    st.session_state.log_busca = []


# ── Busca (executa fora das tabs para rodar independente da tab ativa) ─────────
if buscar_btn:
    hoje = datetime.now()
    data_ini = (hoje - timedelta(days=dias)).strftime("%Y%m%d")
    data_fim = hoje.strftime("%Y%m%d")
    modalidades_ativas = {c: n for c, n in MODALIDADES.items() if mods_sel.get(c)}

    if not modalidades_ativas:
        st.warning("Selecione ao menos uma modalidade.")
        st.stop()

    candidatas = []
    log = []
    total_mods = len(modalidades_ativas)

    with st.status(
        f"Buscando de {data_ini[6:]}/{data_ini[4:6]}/{data_ini[:4]} "
        f"a {data_fim[6:]}/{data_fim[4:6]}/{data_fim[:4]}...",
        expanded=True,
    ) as status_box:
        prog = st.progress(0)

        MAX_PAG_CLOUD = 10  # ~500 registros por modalidade — evita timeout na nuvem
        for i, (cod, nome) in enumerate(modalidades_ativas.items()):
            st.write(f"Consultando **{nome}**...")
            registros = consultar_pncp(data_ini, data_fim, cod, max_paginas=MAX_PAG_CLOUD)
            atingiu_limite = len(registros) >= MAX_PAG_CLOUD * 50
            msg = f"{nome}: {len(registros)} registros" + (" ⚠️ limite atingido — reduza o período" if atingiu_limite else "")
            log.append(msg)
            st.write(f"✅ {msg}")

            for item in registros:
                objeto = item.get("objetoCompra", "")
                palavra = contem_palavra_chave(objeto, st.session_state.get("palavras_ativas"))
                if palavra:
                    valor = item.get("valorTotalEstimado") or 0
                    if valor == 0 or valor >= valor_min:
                        candidatas.append(normalizar_linha(item, nome, palavra))

            prog.progress((i + 1) / total_mods)

        vistas: set = set()
        unicas = []
        for c in candidatas:
            k = c["numeroControlePNCP"]
            if k not in vistas:
                vistas.add(k)
                unicas.append(c)

        prog.empty()
        status_box.update(
            label=f"✅ Concluído — {len(unicas)} licitações encontradas",
            state="complete",
        )

    st.session_state.resultados = unicas
    st.session_state.ultima_busca = datetime.now().strftime("%d/%m/%Y %H:%M")
    st.session_state.log_busca = log


# ── Resultados ────────────────────────────────────────────────────────────────
if st.session_state.resultados:
    resultados = st.session_state.resultados
    df = pd.DataFrame(resultados)

    col1, col2, col3, col4 = st.columns(4)
    com_valor = df[df["valorEstimado"].apply(
        lambda x: isinstance(x, (int, float)) and x > 0
    )]
    with col1:
        st.metric("Total encontrado", len(df))
    with col2:
        st.metric("Com valor informado", len(com_valor))
    with col3:
        if len(com_valor) > 0:
            media = com_valor["valorEstimado"].mean()
            st.metric("Valor médio", f"R$ {media / 1e6:.1f} M")
        else:
            st.metric("Valor médio", "—")
    with col4:
        st.metric("Atualizado em", st.session_state.ultima_busca or "—")

    st.divider()

    df_vis = df.copy()
    if uf_filtro_sidebar:
        df_vis = df_vis[df_vis["uf"].isin(uf_filtro_sidebar)]

    def fmt_valor(v):
        if isinstance(v, (int, float)) and v > 0:
            return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return "—"

    df_vis = df_vis.copy()
    df_vis["valorEstimado"] = df_vis["valorEstimado"].apply(fmt_valor)

    colunas_exibir = [
        "numeroControlePNCP", "uf", "municipio", "orgao", "modalidade",
        "objetoCompra", "valorEstimado", "dataEncerramentoProposta",
        "situacao", "palavraChaveEncontrada", "linkPNCP",
    ]

    st.dataframe(
        df_vis[colunas_exibir],
        use_container_width=True,
        hide_index=True,
        height=520,
        column_config={
            "numeroControlePNCP":        st.column_config.TextColumn("Nº PNCP", width="medium"),
            "uf":                        st.column_config.TextColumn("UF", width="small"),
            "municipio":                 st.column_config.TextColumn("Município", width="medium"),
            "orgao":                     st.column_config.TextColumn("Órgão", width="large"),
            "modalidade":                st.column_config.TextColumn("Modalidade", width="medium"),
            "objetoCompra":              st.column_config.TextColumn("Objeto", width="large"),
            "valorEstimado":             st.column_config.TextColumn("Valor Estimado", width="medium"),
            "dataEncerramentoProposta":  st.column_config.TextColumn("Encerramento Proposta", width="medium"),
            "situacao":                  st.column_config.TextColumn("Situação", width="medium"),
            "palavraChaveEncontrada":    st.column_config.TextColumn("Palavra-chave", width="medium"),
            "linkPNCP":                  st.column_config.LinkColumn("Link PNCP", width="medium"),
        },
    )

    def gerar_excel(dados: list) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "Licitacoes"
        ws.append(CABECALHO)
        fonte_cab = Font(bold=True, color="FFFFFF", name="Arial")
        fill_cab = PatternFill("solid", start_color="1F4E78")
        for col_idx in range(1, len(CABECALHO) + 1):
            c = ws.cell(row=1, column=col_idx)
            c.font = fonte_cab
            c.fill = fill_cab
            c.alignment = Alignment(horizontal="center", vertical="center")
        for linha in dados:
            ws.append([sanitizar(linha.get(col, "")) for col in CABECALHO])
        larguras = [34, 16, 18, 32, 16, 6, 20, 22, 50, 16, 18, 18, 18, 16, 20, 40, 20]
        for col_idx, larg in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = larg
        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    col_dl, col_info = st.columns([1, 4])
    with col_dl:
        excel_buf = gerar_excel(resultados)
        st.download_button(
            label="⬇️  Baixar Excel",
            data=excel_buf,
            file_name=f"licitacoes_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
    with col_info:
        st.caption(f"Exporta todas as {len(resultados)} licitações (independente do filtro de UF).")


    with st.expander("📋 Log da última busca"):
        for entry in st.session_state.log_busca:
            st.caption(f"• {entry}")

else:
    st.info(
        "👈 Configure os parâmetros na barra lateral e clique em **Buscar Licitações** para começar.",
        icon="ℹ️",
    )
    with st.expander("Como usar"):
        st.markdown("""
1. **Período**: quantos dias retroativos consultar no PNCP (padrão: 9 dias)
2. **Valor mínimo**: obras abaixo desse valor são descartadas (padrão: R$ 100 M)
3. **Modalidades**: marque os tipos de licitação desejados
4. Clique em **Buscar** — a busca pode levar alguns minutos dependendo do volume
5. Filtre por estado na tabela e use **Baixar Excel** para exportar
        """)
