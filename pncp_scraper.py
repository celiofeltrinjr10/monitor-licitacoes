"""
Monitor semanal de licitações de OBRAS no PNCP (Portal Nacional de Contratações Públicas)

O que este script faz:
1. Consulta a API pública do PNCP (sem necessidade de login/senha) por um período de datas.
2. Varre as modalidades de contratação relevantes para obras (Concorrência e Pregão).
3. Filtra os resultados pelo OBJETO da contratação usando uma lista de palavras-chave de obras.
4. Compara com o que já existe na planilha de controle (chave única: numeroControlePNCP).
5. Acrescenta SOMENTE as linhas novas na planilha, sem duplicar nem sobrescrever histórico.

Como rodar:
    python pncp_scraper.py                  -> busca os últimos 9 dias (padrão, com folga de segurança)
    python pncp_scraper.py --dias 14         -> busca os últimos 14 dias
    python pncp_scraper.py --dry-run         -> mostra o que encontraria, sem salvar na planilha

Agendamento: ver README.md na mesma pasta.
"""

import argparse
import sys
import time
import unicodedata
from datetime import datetime, timedelta
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError
import re

# Caracteres de controle que o formato XLSX não aceita em células de texto
_CARACTERES_ILEGAIS = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f]"
)

# ----------------------------------------------------------------------------
# CONFIGURAÇÃO — ajuste livremente conforme sua necessidade
# ----------------------------------------------------------------------------

PASTA = Path(__file__).parent
PLANILHA = PASTA / "controle_licitacoes_obras.xlsx"
LOG_ARQUIVO = PASTA / "log_execucoes.txt"

BASE_URL = "https://pncp.gov.br/api/consulta/v1/contratacoes/publicacao"

# Modalidades relevantes para obras de construção civil (tabela de domínio do PNCP).
# 4 = Concorrência Eletrônica | 5 = Concorrência Presencial
# 6 = Pregão Eletrônico       | 7 = Pregão Presencial
# Pregão entra na lista porque obras de menor porte / serviços de engenharia
# às vezes saem como pregão. Dispensa (8) fica de fora por padrão (alto volume
# de ruído) — descomente abaixo se quiser incluir.
MODALIDADES = {
    4: "Concorrência Eletrônica",
    5: "Concorrência Presencial",
    6: "Pregão Eletrônico",
    7: "Pregão Presencial",
    # 8: "Dispensa de Licitação",
}

# Lista de palavras-chave alinhada ao escopo da S.A. Paulista:
# construção pesada, infraestrutura viária, mobilidade urbana, hidráulica e energia.
# O filtro é aplicado sobre o campo "objetoCompra", sem distinção de acentos
# ou maiúsculas/minúsculas. Edite esta lista a qualquer momento.
PALAVRAS_CHAVE_OBRAS = [
    # --- Infraestrutura Rodoviária ---
    "rodovia", "duplicacao", "restauracao rodoviaria", "implantacao de rodovia",
    "pavimentacao asfáltica", "recapeamento", "terraplanagem",
    "obras de arte especiais", "obra de arte especial",
    "contorno viario", "anel viario", "acesso rodoviario",

    # --- Vias Urbanas e Mobilidade ---
    "complexo viario", "sistema viario", "via expressa",
    "corredor de onibus", "corredor brt", "corredor exclusivo",
    "rotatoria", "trincheira", "passagem subterranea",
    "intersecao viaria", "adequacao viaria",

    # --- Estruturas Especiais ---
    "viaduto", "ponte", "tunel",

    # --- Metrô, Monotrilho e Trilhos ---
    "sistema metro", "monotrilho", "vlt",
    "ferrovia", "obra ferroviaria", "linha ferroviaria",

    # --- Canalizações e Infraestrutura Hídrica ---
    "canalizacao", "transposicao de rio", "canal de aducao",
    "adutora", "sifao", "estacao de bombeamento", "obra hidrica",

    # --- Barragens e Hidrelétricas ---
    "barragem", "usina hidreletrica", "central hidreletrica",

    # --- Aeroportos ---
    "aeroporto", "pista de pouso", "patio de aeronaves",

    # --- Geral Construção Pesada ---
    "obras civis", "obra de engenharia",
    "infraestrutura viaria", "infraestrutura de transporte",
    "drenagem", "pavimentacao",
]

TAMANHO_PAGINA = 50
MAX_PAGINAS_SEGURANCA = 200  # trava de segurança contra loop infinito
PAUSA_ENTRE_REQUISICOES = 1.0  # segundos, para não sobrecarregar a API pública
MAX_TENTATIVAS_429 = 5  # quantas vezes tenta de novo se a API responder "muitas requisições"

# Filtro de valor mínimo estimado (R$).
# Obras abaixo deste valor são descartadas.
# Exceção: valor == 0 ou nulo passa (pode ser RDCI ou orçamento sigiloso — vale conferir).
VALOR_MINIMO = 100_000_000  # R$ 100 milhões

CABECALHO = [
    "numeroControlePNCP", "dataColeta", "statusInterno", "orgao", "cnpjOrgao",
    "uf", "municipio", "modalidade", "objetoCompra", "valorEstimado",
    "dataPublicacaoPNCP", "dataAberturaProposta", "dataEncerramentoProposta",
    "situacao", "palavraChaveEncontrada", "linkPNCP", "processo",
]


def remover_acentos(texto: str) -> str:
    if not texto:
        return ""
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower()


def sanitizar(valor):
    """Remove caracteres de controle que o XLSX não aceita. Não altera números/None."""
    if isinstance(valor, str):
        return _CARACTERES_ILEGAIS.sub("", valor)
    return valor


def contem_palavra_chave(objeto: str, lista: list | None = None) -> str | None:
    objeto_normalizado = remover_acentos(objeto)
    palavras = lista if lista is not None else PALAVRAS_CHAVE_OBRAS
    for palavra in palavras:
        if remover_acentos(palavra) in objeto_normalizado:
            return palavra
    return None


def consultar_pncp(data_inicial: str, data_final: str, modalidade_id: int,
                   max_paginas: int = MAX_PAGINAS_SEGURANCA) -> list[dict]:
    """Pagina sobre a API de consulta do PNCP para uma modalidade e devolve todos os registros."""
    resultados = []
    pagina = 1
    pausa_atual = PAUSA_ENTRE_REQUISICOES  # aumenta dinamicamente após 429, recupera gradualmente
    while pagina <= min(max_paginas, MAX_PAGINAS_SEGURANCA):
        params = {
            "dataInicial": data_inicial,
            "dataFinal": data_final,
            "codigoModalidadeContratacao": modalidade_id,
            "pagina": pagina,
            "tamanhoPagina": TAMANHO_PAGINA,
        }

        dados = None
        houve_429 = False
        for tentativa in range(1, MAX_TENTATIVAS_429 + 1):
            try:
                resp = requests.get(BASE_URL, params=params, timeout=30)
                if resp.status_code == 429:
                    houve_429 = True
                    espera = float(resp.headers.get("Retry-After", tentativa * 5))
                    print(f"  [aviso] 429 recebido (tentativa {tentativa}/{MAX_TENTATIVAS_429}). "
                          f"Aguardando {espera:.0f}s antes de tentar novamente...")
                    time.sleep(espera)
                    continue
                resp.raise_for_status()
                payload = resp.json()
                dados = payload.get("data", [])
                break
            except requests.RequestException as exc:
                print(f"  [erro] modalidade={modalidade_id} pagina={pagina}: {exc}")
                break

        if dados is None:
            # esgotou as tentativas ou deu erro não relacionado a rate limit: para esta modalidade
            break
        if not dados:
            break

        resultados.extend(dados)

        # Backoff adaptativo: dobra a pausa após 429 (máx 10s), recupera 10% por página limpa
        if houve_429:
            pausa_atual = min(pausa_atual * 2, 10.0)
        else:
            pausa_atual = max(pausa_atual * 0.9, PAUSA_ENTRE_REQUISICOES)

        total_paginas = payload.get("totalPaginas")
        if total_paginas is not None and pagina >= total_paginas:
            break

        pagina += 1
        time.sleep(pausa_atual)

    return resultados


def montar_link_pncp(item: dict) -> str:
    cnpj = item.get("orgaoEntidade", {}).get("cnpj", "")
    ano = item.get("anoCompra", "")
    sequencial = item.get("sequencialCompra", "")
    if cnpj and ano and sequencial:
        return f"https://pncp.gov.br/app/editais/{cnpj}/{ano}/{sequencial}"
    return ""


def normalizar_linha(item: dict, modalidade_nome: str, palavra_encontrada: str) -> dict:
    orgao = item.get("orgaoEntidade", {}) or {}
    unidade = item.get("unidadeOrgao", {}) or {}
    return {
        "numeroControlePNCP": item.get("numeroControlePNCP", ""),
        "dataColeta": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "statusInterno": "novo - a analisar",
        "orgao": orgao.get("razaoSocial", ""),
        "cnpjOrgao": orgao.get("cnpj", ""),
        "uf": unidade.get("ufSigla", ""),
        "municipio": unidade.get("municipioNome", ""),
        "modalidade": modalidade_nome,
        "objetoCompra": item.get("objetoCompra", ""),
        "valorEstimado": item.get("valorTotalEstimado", ""),
        "dataPublicacaoPNCP": item.get("dataPublicacaoPncp", ""),
        "dataAberturaProposta": item.get("dataAberturaProposta", ""),
        "dataEncerramentoProposta": item.get("dataEncerramentoProposta", ""),
        "situacao": item.get("situacaoCompraNome", ""),
        "palavraChaveEncontrada": palavra_encontrada,
        "linkPNCP": montar_link_pncp(item),
        "processo": item.get("processo", ""),
    }


def criar_planilha_se_nao_existir():
    if PLANILHA.exists():
        return
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Licitacoes"
    sheet.append(CABECALHO)

    fonte_cabecalho = Font(bold=True, color="FFFFFF", name="Arial")
    preenchimento = PatternFill("solid", start_color="1F4E78")
    for col_idx, _ in enumerate(CABECALHO, start=1):
        celula = sheet.cell(row=1, column=col_idx)
        celula.font = fonte_cabecalho
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal="center", vertical="center")

    larguras = [34, 16, 18, 32, 16, 6, 20, 22, 50, 16, 18, 18, 18, 16, 20, 40, 20]
    for col_idx, largura in enumerate(larguras, start=1):
        sheet.column_dimensions[get_column_letter(col_idx)].width = largura

    sheet.freeze_panes = "A2"
    wb.save(PLANILHA)
    print(f"Planilha criada em: {PLANILHA}")


def carregar_chaves_existentes() -> set[str]:
    wb = load_workbook(PLANILHA, read_only=True)
    sheet = wb["Licitacoes"]
    chaves = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            chaves.add(str(row[0]))
    wb.close()
    return chaves


def acrescentar_linhas(linhas_novas: list[dict]):
    wb = load_workbook(PLANILHA)
    sheet = wb["Licitacoes"]
    ignoradas = 0
    for linha in linhas_novas:
        valores = [sanitizar(linha.get(col, "")) for col in CABECALHO]
        try:
            sheet.append(valores)
        except IllegalCharacterError:
            # segurança extra: se ainda assim algum caractere escapar, salva sem ele
            valores_forcados = [str(v).encode("ascii", "ignore").decode() if isinstance(v, str) else v
                                 for v in valores]
            sheet.append(valores_forcados)
            ignoradas += 1
    wb.save(PLANILHA)
    if ignoradas:
        print(f"  [aviso] {ignoradas} linha(s) tiveram caracteres especiais removidos por segurança.")


def registrar_log(mensagem: str):
    with open(LOG_ARQUIVO, "a", encoding="utf-8") as f:
        f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {mensagem}\n")


def main():
    parser = argparse.ArgumentParser(description="Monitor semanal de licitações de obras no PNCP")
    parser.add_argument("--dias", type=int, default=9,
                         help="Quantidade de dias para olhar para trás (padrão: 9, com folga de segurança)")
    parser.add_argument("--dry-run", action="store_true",
                         help="Não salva na planilha, só mostra o que encontraria")
    args = parser.parse_args()

    hoje = datetime.now()
    data_inicial = (hoje - timedelta(days=args.dias)).strftime("%Y%m%d")
    data_final = hoje.strftime("%Y%m%d")

    print(f"Buscando contratações publicadas entre {data_inicial} e {data_final}...")

    criar_planilha_se_nao_existir()
    chaves_existentes = carregar_chaves_existentes()
    print(f"Planilha atual já contém {len(chaves_existentes)} registros.")

    candidatas = []
    for modalidade_id, modalidade_nome in MODALIDADES.items():
        print(f"Consultando modalidade: {modalidade_nome} (codigo {modalidade_id})...")
        registros = consultar_pncp(data_inicial, data_final, modalidade_id)
        print(f"  -> {len(registros)} registros retornados pela API.")
        for item in registros:
            objeto = item.get("objetoCompra", "")
            palavra = contem_palavra_chave(objeto)
            if palavra:
                valor = item.get("valorTotalEstimado") or 0
                if valor == 0 or valor >= VALOR_MINIMO:
                    candidatas.append(normalizar_linha(item, modalidade_nome, palavra))

    print(f"Total de candidatas após filtro de palavras-chave: {len(candidatas)}")

    novas = [c for c in candidatas if c["numeroControlePNCP"] not in chaves_existentes]
    # remove duplicatas dentro do próprio lote (caso uma licitação apareça em duas modalidades)
    vistas = set()
    novas_unicas = []
    for n in novas:
        if n["numeroControlePNCP"] not in vistas:
            vistas.add(n["numeroControlePNCP"])
            novas_unicas.append(n)

    print(f"Linhas realmente novas (não duplicadas na planilha): {len(novas_unicas)}")

    if args.dry_run:
        for n in novas_unicas:
            print(f"  [{n['uf']}] {n['orgao']} - {n['objetoCompra'][:80]}")
        print("Modo --dry-run: nada foi salvo na planilha.")
        return

    if novas_unicas:
        acrescentar_linhas(novas_unicas)
        print(f"{len(novas_unicas)} novas licitações adicionadas à planilha.")
    else:
        print("Nenhuma licitação nova encontrada nesta execução.")

    registrar_log(
        f"dias={args.dias} candidatas={len(candidatas)} novas={len(novas_unicas)} "
        f"total_planilha={len(chaves_existentes) + len(novas_unicas)}"
    )


if __name__ == "__main__":
    sys.exit(main())
